from __future__ import annotations

import argparse
import math
import re
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from pymongo import MongoClient


DEFAULT_MONGO_URI = "mongodb://127.0.0.1:27017/"
DEFAULT_DB = "pdd_sales_trends"
DEFAULT_OUT_DIR = Path(r"E:\proj\MyShop\output")
DEFAULT_DATA_ROOT = Path(r"E:\ds电商\data")
TARGET_NET_PROFIT = 20_000.0
PLATFORM_FEE_RATE = 0.006


@dataclass(frozen=True)
class ProductDirection:
    name: str
    keywords: tuple[str, ...]
    exclude_keywords: tuple[str, ...] = ()
    positioning: str = ""
    opportunity: str = ""
    risk: str = ""
    role: str = "主推"
    shipping_pack: float = 10.0
    ad_cpa: float = 10.0
    aftersale_rate: float = 0.03
    max_sku_cost: float | None = None


DIRECTIONS: tuple[ProductDirection, ...] = (
    ProductDirection(
        name="室外路冲补角石",
        keywords=("路冲", "补角", "缺角", "屋角", "室外", "门口", "大门口", "化煞"),
        positioning="入户门、墙角、路冲、庭院场景款，SKU 重点覆盖 24-60cm 石板/原石。",
        opportunity="购买目的明确，搜索词和标题高频，适合用场景主图承接转化。",
        risk="中大件破损和运费风险高，详情页必须写清尺寸、签收和天然色差。",
        shipping_pack=18,
        ad_cpa=16,
        max_sku_cost=220,
    ),
    ProductDirection(
        name="靠山石/办公室摆件",
        keywords=("靠山石", "办公室", "老板桌", "办公桌", "收银台", "客厅"),
        positioning="办公桌、老板桌、收银台、送礼场景，中小尺寸带底座款优先。",
        opportunity="办公/送礼文案容易聚焦，客单可高于普通原石小摆件。",
        risk="功效型文案要克制，避免绝对化承诺；随机原石需提前说明。",
        shipping_pack=12,
        ad_cpa=12,
        max_sku_cost=130,
    ),
    ProductDirection(
        name="石板刻字门牌",
        keywords=("门牌", "刻字", "石板", "定制", "外墙", "村牌"),
        positioning="室外门牌、刻字定制、院门/店门场景，做利润款。",
        opportunity="定制属性提升客单和比价难度，适合承接高意向用户。",
        risk="客服沟通、刻字确认和发货周期成本高，必须设置下单备注流程。",
        role="利润款",
        shipping_pack=22,
        ad_cpa=18,
        aftersale_rate=0.04,
        max_sku_cost=320,
    ),
    ProductDirection(
        name="青石雕刻石敢当",
        keywords=("青石", "雕刻", "浮雕", "阳刻", "阴刻"),
        exclude_keywords=("挂件", "挂饰", "五帝钱", "中国结"),
        positioning="青石材质、雕刻清晰、室外耐放，中高客单利润款。",
        opportunity="材质和工艺有差异化，主图质感更稳定。",
        risk="重量和破损风险高，低价 SKU 不宜承诺同图同款。",
        role="利润款",
        shipping_pack=24,
        ad_cpa=20,
        aftersale_rate=0.04,
        max_sku_cost=360,
    ),
    ProductDirection(
        name="挂件/中式挂饰",
        keywords=("挂件", "挂饰", "五帝钱", "牌匾", "桃木", "中国结"),
        positioning="小件挂饰、乔迁礼品、搭配加购款，控制低物流成本。",
        opportunity="体积小、发货压力低，可作为店铺动销和关联搭配。",
        risk="与石敢当主词相关性弱，竞争词偏泛，不宜重投广告。",
        role="补充款",
        shipping_pack=6,
        ad_cpa=8,
        max_sku_cost=80,
    ),
    ProductDirection(
        name="大号庭院镇宅石",
        keywords=("庭院", "大号", "大型", "景观", "园林", "户外", "企业门牌"),
        positioning="庭院、园林、户外景观高客单款，只做少量询单利润款。",
        opportunity="客单高，单笔利润空间大，可用于拉高店铺价格带。",
        risk="物流、木架、签收、售后风险最高，首月不要压库存。",
        role="高客单测试",
        shipping_pack=90,
        ad_cpa=35,
        aftersale_rate=0.05,
        max_sku_cost=900,
    ),
    ProductDirection(
        name="一图一物精选原石",
        keywords=("一图一物", "精选", "精品", "奇石"),
        positioning="一石一拍、客服确认后发货，做差异化精选款。",
        opportunity="减少同质化比价，适合用真实图片提高信任。",
        risk="上新维护成本高，代发需要确认源商可按图发货。",
        role="差异化款",
        shipping_pack=14,
        ad_cpa=14,
        max_sku_cost=180,
    ),
    ProductDirection(
        name="天然原石小摆件",
        keywords=("原石", "天然", "石头", "奇石"),
        exclude_keywords=("大号", "大型", "庭院", "景观", "门牌", "刻字"),
        positioning="9.9-39.9 引流款，覆盖桌面、门口、小摆件需求。",
        opportunity="搜索覆盖广，低客单更适合拼多多测款和引流。",
        risk="差异化弱，广告必须严格限额，利润依赖 SKU 梯度升级。",
        role="引流款",
        shipping_pack=8,
        ad_cpa=6,
        max_sku_cost=70,
    ),
    ProductDirection(
        name="金属招财衍生款",
        keywords=("金属", "招财", "天官", "铜", "福星"),
        positioning="关联搭配和扩品款，不抢主推预算。",
        opportunity="可丰富店铺货盘，承接部分招财/挂件需求。",
        risk="与泰山石主词相关性弱，不能作为首月利润核心。",
        role="关联款",
        shipping_pack=8,
        ad_cpa=8,
        max_sku_cost=120,
    ),
    ProductDirection(
        name="普通泰山石摆件/朱砂红石测试款",
        keywords=("泰山石", "石敢当", "朱砂", "红石", "红色"),
        positioning="基础款和视觉差异款，小流量测试主图与价格。",
        opportunity="覆盖面广，可补足店铺基础 SKU 和红色差异化素材。",
        risk="同质化强，必须靠主图、价格带和评价素材建立信任。",
        role="测试款",
        shipping_pack=10,
        ad_cpa=8,
        max_sku_cost=90,
    ),
)


def sales_num(value: Any) -> int:
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        return int(value)
    digits = re.sub(r"[^\d]", "", str(value))
    return int(digits or 0)


def contains_any(text: str, keywords: tuple[str, ...]) -> bool:
    return any(keyword in text for keyword in keywords)


def normalize(value: float, max_value: float) -> float:
    if max_value <= 0:
        return 0.0
    return max(0.0, min(value / max_value, 1.0))


def suggest_price(cost: float) -> float:
    if cost < 20:
        raw = cost * 1.75 + 4
    elif cost < 50:
        raw = cost * 1.58 + 6
    elif cost < 100:
        raw = cost * 1.45 + 10
    elif cost < 300:
        raw = cost * 1.35 + 18
    else:
        raw = cost * 1.30 + 38
    return round(math.ceil(raw) - 0.1, 1)


def estimate_financials(direction: ProductDirection, cost: float) -> dict[str, float]:
    price = suggest_price(cost)
    platform_fee = round(price * PLATFORM_FEE_RATE, 2)
    aftersale = round(max(price * direction.aftersale_rate, 1.5), 2)
    net_profit = round(price - cost - direction.shipping_pack - platform_fee - direction.ad_cpa - aftersale, 2)

    if net_profit < 8 and cost < 80:
        price = round(math.ceil((cost + direction.shipping_pack + direction.ad_cpa + aftersale + 8) / (1 - PLATFORM_FEE_RATE)) - 0.1, 1)
        platform_fee = round(price * PLATFORM_FEE_RATE, 2)
        aftersale = round(max(price * direction.aftersale_rate, 1.5), 2)
        net_profit = round(price - cost - direction.shipping_pack - platform_fee - direction.ad_cpa - aftersale, 2)

    return {
        "suggest_price": price,
        "shipping_pack": direction.shipping_pack,
        "platform_fee": platform_fee,
        "ad_cpa": direction.ad_cpa,
        "aftersale_reserve": aftersale,
        "net_profit": net_profit,
        "ad_stop_line": round(max(direction.ad_cpa * 1.35, net_profit * 0.55), 2),
    }


def median(values: list[float]) -> float | None:
    if not values:
        return None
    ordered = sorted(values)
    mid = len(ordered) // 2
    if len(ordered) % 2:
        return ordered[mid]
    return round((ordered[mid - 1] + ordered[mid]) / 2, 2)


def safe_round(value: Any, digits: int = 2) -> Any:
    if value is None:
        return None
    return round(float(value), digits)


def load_runs(db: Any) -> tuple[dict[str, Any], dict[str, Any]]:
    runs = list(db.crawl_runs.find({"keyword": "泰山 石敢当"}).sort("crawl_time", 1))
    if len(runs) < 2:
        raise RuntimeError("MongoDB 中至少需要 2 次“泰山 石敢当”抓取记录，才能计算趋势。")
    return runs[0], runs[-1]


def choose_sku(direction: ProductDirection, sku_docs: list[dict[str, Any]]) -> dict[str, Any] | None:
    valid = [
        sku
        for sku in sku_docs
        if sku.get("wholesale_price_yuan") is not None
        and sku.get("wholesale_price_yuan") > 0
        and (direction.max_sku_cost is None or sku.get("wholesale_price_yuan") <= direction.max_sku_cost)
    ]
    if not valid and direction.max_sku_cost is None:
        valid = [sku for sku in sku_docs if sku.get("wholesale_price_yuan") is not None and sku.get("wholesale_price_yuan") > 0]
    if not valid:
        return None

    target_cost = {
        "引流款": 18,
        "主推": 45,
        "利润款": 120,
        "高客单测试": 380,
        "差异化款": 60,
        "补充款": 35,
        "关联款": 35,
        "测试款": 28,
    }.get(direction.role, 45)

    def sku_score(sku: dict[str, Any]) -> float:
        cost = float(sku["wholesale_price_yuan"])
        fin = estimate_financials(direction, cost)
        spec = str(sku.get("specs") or "")
        size_bonus = 8 if any(k in spec for k in ("30", "40", "带底座", "阳刻", "八卦")) else 0
        return fin["net_profit"] * 5 + size_bonus - abs(cost - target_cost) * 0.25

    return sorted(valid, key=sku_score, reverse=True)[0]


def build_direction_rows(
    latest_docs: list[dict[str, Any]],
    first_by_goods: dict[str, dict[str, Any]],
    skus_by_goods: dict[str, list[dict[str, Any]]],
) -> tuple[pd.DataFrame, pd.DataFrame]:
    rows: list[dict[str, Any]] = []
    sku_rows: list[dict[str, Any]] = []
    used_goods: set[str] = set()

    max_sales = max((sales_num(doc.get("sales_tip_amount")) for doc in latest_docs), default=1)
    max_delta = max(
        (
            max(0, sales_num(doc.get("sales_tip_amount")) - sales_num(first_by_goods.get(doc["goods_id"], {}).get("sales_tip_amount")))
            for doc in latest_docs
        ),
        default=1,
    )

    for direction in DIRECTIONS:
        candidates: list[dict[str, Any]] = []
        for doc in latest_docs:
            title = str(doc.get("title") or "")
            if doc["goods_id"] in used_goods:
                continue
            if not contains_any(title, direction.keywords):
                continue
            if direction.exclude_keywords and contains_any(title, direction.exclude_keywords):
                continue
            if doc.get("min_wholesale_price_yuan") is None or not skus_by_goods.get(doc["goods_id"]):
                continue
            candidates.append(doc)

        if not candidates:
            continue

        product_count = len(candidates)
        total_sales = sum(sales_num(doc.get("sales_tip_amount")) for doc in candidates)
        total_delta = sum(
            max(0, sales_num(doc.get("sales_tip_amount")) - sales_num(first_by_goods.get(doc["goods_id"], {}).get("sales_tip_amount")))
            for doc in candidates
        )
        min_prices = [float(doc["min_wholesale_price_yuan"]) for doc in candidates if doc.get("min_wholesale_price_yuan") is not None]
        sku_counts = [int(doc.get("sku_count") or 0) for doc in candidates]
        detail_counts = [int(len(doc.get("property_texts") or [])) for doc in candidates]

        best_doc: dict[str, Any] | None = None
        best_sku: dict[str, Any] | None = None
        best_fin: dict[str, float] | None = None
        best_score = -999.0

        for doc in candidates:
            sku = choose_sku(direction, skus_by_goods.get(doc["goods_id"], []))
            if not sku:
                continue
            cost = float(sku["wholesale_price_yuan"])
            fin = estimate_financials(direction, cost)
            current_sales = sales_num(doc.get("sales_tip_amount"))
            previous_sales = sales_num(first_by_goods.get(doc["goods_id"], {}).get("sales_tip_amount"))
            delta = max(0, current_sales - previous_sales)
            sku_count = int(doc.get("sku_count") or 0)
            detail_count = len(doc.get("property_texts") or [])
            risk_penalty = 0.18 if direction.role == "高客单测试" else 0.08 if direction.role in {"利润款", "差异化款"} else 0.03
            score = (
                normalize(delta, max_delta) * 30
                + normalize(current_sales, max_sales) * 15
                + normalize(max(fin["net_profit"], 0), 120) * 25
                + normalize(sku_count, 30) * 10
                + normalize(detail_count, 16) * 10
                + (1 - risk_penalty) * 10
            )
            if score > best_score:
                best_doc = doc
                best_sku = sku
                best_fin = fin
                best_score = score

        if not best_doc or not best_sku or not best_fin:
            continue

        goods_id = best_doc["goods_id"]
        used_goods.add(goods_id)
        current_sales = sales_num(best_doc.get("sales_tip_amount"))
        previous_sales = sales_num(first_by_goods.get(goods_id, {}).get("sales_tip_amount"))
        sales_delta = max(0, current_sales - previous_sales)
        target_orders = math.ceil(TARGET_NET_PROFIT / max(best_fin["net_profit"], 1))

        row = {
            "排名": len(rows) + 1,
            "产品方向": direction.name,
            "货盘角色": direction.role,
            "综合分": round(best_score, 1),
            "商品池数量": product_count,
            "商品池累计销量": total_sales,
            "商品池区间销量增量": total_delta,
            "源价中位数": safe_round(median(min_prices)),
            "SKU中位数": safe_round(median([float(x) for x in sku_counts]), 1),
            "代表源商品ID": goods_id,
            "代表标题": best_doc.get("title") or "",
            "代表当前累计销量": current_sales,
            "代表区间销量增量": sales_delta,
            "建议SKU": best_sku.get("specs") or "",
            "拿货价": safe_round(best_sku.get("wholesale_price_yuan")),
            "建议售价": best_fin["suggest_price"],
            "估算运费包装": best_fin["shipping_pack"],
            "平台扣费": best_fin["platform_fee"],
            "广告获客成本": best_fin["ad_cpa"],
            "售后损耗预留": best_fin["aftersale_reserve"],
            "预估单件净利": best_fin["net_profit"],
            "月目标订单数": target_orders,
            "广告止损线/单": best_fin["ad_stop_line"],
            "机会理由": direction.opportunity,
            "建议定位": direction.positioning,
            "履约风险": direction.risk,
            "图片字段": best_doc.get("image_local_file") or "",
            "店铺/商家": best_doc.get("shop_name") or best_doc.get("mall") or "",
        }
        rows.append(row)

        sorted_skus = sorted(
            [sku for sku in skus_by_goods.get(goods_id, []) if sku.get("wholesale_price_yuan") is not None],
            key=lambda sku: sku["wholesale_price_yuan"],
        )[:8]
        for sku in sorted_skus:
            fin = estimate_financials(direction, float(sku["wholesale_price_yuan"]))
            sku_rows.append(
                {
                    "产品方向": direction.name,
                    "源商品ID": goods_id,
                    "SKU ID": sku.get("sku_id") or "",
                    "SKU规格": sku.get("specs") or "",
                    "拿货价": safe_round(sku.get("wholesale_price_yuan")),
                    "建议售价": fin["suggest_price"],
                    "预估单件净利": fin["net_profit"],
                    "库存": sku.get("quantity"),
                    "缩略图": sku.get("thumb_local_file") or sku.get("thumb_url") or "",
                }
            )

    result = pd.DataFrame(rows).sort_values(["综合分", "预估单件净利"], ascending=False).reset_index(drop=True)
    result["排名"] = range(1, len(result) + 1)
    return result, pd.DataFrame(sku_rows)


def build_scenario_rows(selection: pd.DataFrame) -> pd.DataFrame:
    main = selection.copy()
    main["权重"] = main["货盘角色"].map(
        {
            "主推": 1.25,
            "利润款": 1.15,
            "引流款": 0.95,
            "差异化款": 0.8,
            "补充款": 0.55,
            "关联款": 0.45,
            "高客单测试": 0.4,
            "测试款": 0.45,
        }
    ).fillna(0.8)
    main["正净利"] = main["预估单件净利"].clip(lower=1)

    scenarios = [
        ("保守", 0.85, 3000),
        ("中性", 1.00, 5000),
        ("进取", 1.10, 8000),
    ]
    rows: list[dict[str, Any]] = []
    for scenario, profit_factor, fixed_reserve in scenarios:
        main["场景单件净利"] = (main["正净利"] * profit_factor).round(2)
        weighted_profit = (main["场景单件净利"] * main["权重"]).sum()
        order_units = (TARGET_NET_PROFIT + fixed_reserve) / max(weighted_profit, 1)
        for _, row in main.iterrows():
            orders = max(1, math.ceil(order_units * float(row["权重"])))
            gross_net = round(orders * float(row["场景单件净利"]), 2)
            rows.append(
                {
                    "场景": scenario,
                    "产品方向": row["产品方向"],
                    "计划订单数": orders,
                    "单件净利": row["场景单件净利"],
                    "商品净利贡献": gross_net,
                    "固定广告/售后预留池": fixed_reserve,
                }
            )
    scenarios_df = pd.DataFrame(rows)
    return scenarios_df


def build_validation_rows(db: Any, latest_run: str, selection: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for collection in ("crawl_runs", "goods", "goods_snapshots", "sku_snapshots"):
        count = db[collection].count_documents({})
        rows.append({"检查项": f"{collection} 集合数量", "结果": "通过" if count else "失败", "详情": count})

    latest_count = db.goods_snapshots.count_documents({"run_id": latest_run})
    latest_sku_count = db.sku_snapshots.count_documents({"run_id": latest_run})
    rows.extend(
        [
            {"检查项": "最新批次商品快照", "结果": "通过" if latest_count else "失败", "详情": latest_count},
            {"检查项": "最新批次 SKU 快照", "结果": "通过" if latest_sku_count else "失败", "详情": latest_sku_count},
            {"检查项": "选品方向数量", "结果": "通过" if len(selection) >= 10 else "警告", "详情": len(selection)},
            {
                "检查项": "代表商品字段完整",
                "结果": "通过"
                if selection[["代表标题", "建议SKU", "拿货价", "建议售价", "图片字段"]].notna().all().all()
                else "警告",
                "详情": "检查标题、SKU、价格、图片字段",
            },
        ]
    )
    return pd.DataFrame(rows)


def resolve_asset_path(value: Any, data_root: Path) -> Path | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    path = Path(text)
    if path.is_absolute() and path.exists():
        return path
    candidates = [
        data_root / text,
        data_root / text.replace("/", "\\"),
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return None


def embed_main_images(xlsx_path: Path, selection: pd.DataFrame, data_root: Path) -> int:
    workbook = load_workbook(xlsx_path)
    worksheet = workbook["10个上架方向"]
    image_col = 2
    worksheet.insert_cols(image_col)
    worksheet.cell(row=1, column=image_col, value="主图")
    worksheet.column_dimensions["B"].width = 16

    inserted = 0
    for excel_row, (_, row) in enumerate(selection.iterrows(), start=2):
        worksheet.row_dimensions[excel_row].height = 82
        image_path = resolve_asset_path(row.get("图片字段"), data_root)
        if not image_path:
            worksheet.cell(row=excel_row, column=image_col, value="图片缺失")
            continue
        image = OpenpyxlImage(str(image_path))
        image.width = 96
        image.height = 96
        worksheet.add_image(image, f"B{excel_row}")
        inserted += 1

    workbook.save(xlsx_path)
    return inserted


def write_excel_with_fallback(
    xlsx_path: Path,
    selection: pd.DataFrame,
    sku_rows: pd.DataFrame,
    scenarios: pd.DataFrame,
    validation: pd.DataFrame,
) -> Path:
    def write(path: Path) -> None:
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            selection.to_excel(writer, sheet_name="10个上架方向", index=False)
            sku_rows.to_excel(writer, sheet_name="SKU梯度样例", index=False)
            scenarios.to_excel(writer, sheet_name="首月目标拆解", index=False)
            validation.to_excel(writer, sheet_name="校验结果", index=False)

    try:
        write(xlsx_path)
        return xlsx_path
    except PermissionError:
        fallback = xlsx_path.with_name(f"{xlsx_path.stem}_with_images{xlsx_path.suffix}")
        write(fallback)
        return fallback


def write_markdown(
    path: Path,
    selection: pd.DataFrame,
    sku_rows: pd.DataFrame,
    scenarios: pd.DataFrame,
    validation: pd.DataFrame,
    first_run: dict[str, Any],
    latest_run: dict[str, Any],
) -> None:
    scenario_summary = (
        scenarios.groupby("场景")
        .agg(计划订单数=("计划订单数", "sum"), 商品净利贡献=("商品净利贡献", "sum"), 固定广告售后预留=("固定广告/售后预留池", "max"))
        .reset_index()
    )
    scenario_summary["预估扣预留后净利"] = scenario_summary["商品净利贡献"] - scenario_summary["固定广告售后预留"]

    compact_cols = [
        "排名",
        "产品方向",
        "货盘角色",
        "综合分",
        "商品池数量",
        "商品池区间销量增量",
        "代表源商品ID",
        "拿货价",
        "建议售价",
        "预估单件净利",
        "月目标订单数",
        "广告止损线/单",
    ]
    detail_cols = [
        "排名",
        "产品方向",
        "代表标题",
        "建议SKU",
        "机会理由",
        "建议定位",
        "履约风险",
    ]

    lines = [
        "# 泰山石敢当选品与首月 2 万净利润计划",
        "",
        "## 数据口径",
        "",
        f"- MongoDB 数据库：`{DEFAULT_DB}`",
        f"- 趋势区间：`{first_run['_id']}` 到 `{latest_run['_id']}`",
        "- 经营假设：首月目标净利润 20000 元，启动资金 1-3 万，履约方式以代发为主。",
        "- 净利公式：建议售价 - 拿货价 - 估算运费包装 - 平台扣费 - 广告获客成本 - 售后损耗预留。",
        "",
        "## 10 个建议上架方向",
        "",
        selection[compact_cols].to_markdown(index=False),
        "",
        "## 选品说明",
        "",
        selection[detail_cols].to_markdown(index=False),
        "",
        "## 首月目标拆解",
        "",
        scenario_summary.to_markdown(index=False),
        "",
        "## SKU 梯度样例",
        "",
        sku_rows.head(80).to_markdown(index=False),
        "",
        "## 校验结果",
        "",
        validation.to_markdown(index=False),
        "",
        "## 执行重点",
        "",
        "1. 前 7 天只投主推和利润款，单品广告花费超过止损线且无成交就暂停。",
        "2. 引流款不承担利润目标，主要负责拉访客和收藏；利润来自靠山石、刻字门牌、青石雕刻和少量高客单款。",
        "3. 所有天然石 SKU 详情页必须写清随机纹理、色差、尺寸误差、破损签收规则。",
        "4. 大号庭院款首月只接询单和代发，不建议备货。",
    ]
    path.write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate Taishan Shigandang 20k net profit selection plan from MongoDB.")
    parser.add_argument("--mongo-uri", default=DEFAULT_MONGO_URI)
    parser.add_argument("--db", default=DEFAULT_DB)
    parser.add_argument("--out-dir", type=Path, default=DEFAULT_OUT_DIR)
    parser.add_argument("--data-root", type=Path, default=DEFAULT_DATA_ROOT)
    args = parser.parse_args()

    args.out_dir.mkdir(parents=True, exist_ok=True)
    client = MongoClient(args.mongo_uri, serverSelectionTimeoutMS=5000)
    client.admin.command("ping")
    db = client[args.db]

    first_run, latest_run = load_runs(db)
    first_docs = {
        doc["goods_id"]: doc
        for doc in db.goods_snapshots.find({"run_id": first_run["_id"]})
        if doc.get("goods_id")
    }
    latest_docs = list(db.goods_snapshots.find({"run_id": latest_run["_id"], "detail_success": {"$ne": False}}))
    sku_docs = list(db.sku_snapshots.find({"run_id": latest_run["_id"]}))
    skus_by_goods: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for sku in sku_docs:
        if sku.get("goods_id"):
            skus_by_goods[sku["goods_id"]].append(sku)

    selection, sku_rows = build_direction_rows(latest_docs, first_docs, skus_by_goods)
    if len(selection) < 10:
        raise RuntimeError(f"只生成了 {len(selection)} 个方向，少于计划要求的 10 个。")

    scenarios = build_scenario_rows(selection)
    validation = build_validation_rows(db, latest_run["_id"], selection)

    xlsx_path = args.out_dir / "taishan_20k_selection_plan.xlsx"
    md_path = args.out_dir / "taishan_20k_selection_plan.md"

    xlsx_path = write_excel_with_fallback(xlsx_path, selection, sku_rows, scenarios, validation)
    inserted_images = embed_main_images(xlsx_path, selection, args.data_root)
    write_markdown(md_path, selection, sku_rows, scenarios, validation, first_run, latest_run)

    print(f"生成完成: {xlsx_path}")
    print(f"生成完成: {md_path}")
    print(f"已插入主图: {inserted_images} 张")
    print(selection[["排名", "产品方向", "综合分", "代表源商品ID", "建议售价", "预估单件净利", "月目标订单数"]].to_string(index=False))


if __name__ == "__main__":
    main()
